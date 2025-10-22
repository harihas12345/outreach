import json
import os
import io
from typing import Any, Dict, List, Tuple


def _read_excel_from_s3(bucket: str, key: str) -> List[Dict[str, Any]]:
    import boto3  # type: ignore
    from openpyxl import load_workbook  # type: ignore

    s3 = boto3.client("s3")
    obj = s3.get_object(Bucket=bucket, Key=key)
    content = obj["Body"].read()

    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") if h is not None else "" for h in rows[0]]
    required = {"student_id", "student_name", "slack_user_id"}
    if not required.issubset(set(headers)):
        missing = sorted(required - set(headers))
        raise ValueError(f"Missing columns {missing} in {key}")

    id_cols = {"student_id", "student_name", "slack_user_id", "last_active"}
    metric_idxs = [i for i, h in enumerate(headers) if h and h not in id_cols]

    records: List[Dict[str, Any]] = []
    for r in rows[1:]:
        row = list(r) if r is not None else []
        # map by header
        vals = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        metrics: Dict[str, float] = {}
        for i in metric_idxs:
            if i < len(row) and row[i] is not None and str(row[i]).strip() != "":
                try:
                    metrics[headers[i]] = float(row[i])
                except Exception:
                    continue
        last_active_val = vals.get("last_active")
        last_iso = str(last_active_val) if last_active_val not in (None, "") else None
        records.append(
            {
                "studentId": str(vals.get("student_id", "")),
                "studentName": str(vals.get("student_name", "")),
                "slackUserId": str(vals.get("slack_user_id", "")),
                "metrics": metrics,
                "lastActiveIso": last_iso,
            }
        )
    return records


def _compute_deltas(prev: Dict[str, Dict[str, Any]], curr: Dict[str, Dict[str, Any]]):
    deltas: Dict[str, List[Dict[str, Any]]] = {}
    for student_id, c in curr.items():
        p = prev.get(student_id)
        if not p:
            continue
        for metric, cval in c.get("metrics", {}).items():
            pval = p.get("metrics", {}).get(metric)
            if pval is None:
                continue
            change = float(cval) - float(pval)
            deltas.setdefault(student_id, []).append(
                {
                    "studentId": student_id,
                    "metric": metric,
                    "change": change,
                    "previous": float(pval),
                    "current": float(cval),
                }
            )
    return deltas


def _decide_flags(latest_by_student: Dict[str, Dict[str, Any]], deltas_by_student: Dict[str, List[Dict[str, Any]]]):
    from datetime import datetime, timedelta

    flags: Dict[str, List[str]] = {}
    now = datetime.utcnow()
    for sid, rec in latest_by_student.items():
        f: List[str] = []
        last_iso = rec.get("lastActiveIso")
        if last_iso:
            try:
                from datetime import datetime as _dt

                last = _dt.fromisoformat(last_iso)
                if now - last > timedelta(days=7):
                    f.append("inactivity_over_7_days")
            except Exception:
                pass
        else:
            f.append("no_last_active_recorded")

        for d in deltas_by_student.get(sid, []):
            if d.get("change", 0.0) <= -5.0:
                f.append(f"drop_{d['metric']}_{d['previous']}_to_{d['current']}")

        if f:
            flags[sid] = f
    return flags


def _invoke_draft_lambda(student: Dict[str, Any], flags: List[str]) -> str:
    import boto3  # type: ignore

    fn = os.environ.get("AGENT_DRAFT_MESSAGE_LAMBDA_ARN")
    if not fn:
        return _draft_direct(student, flags)
    client = boto3.client("lambda")
    payload = {"student": student, "flags": flags}
    resp = client.invoke(FunctionName=fn, InvocationType="RequestResponse", Payload=json.dumps({"body": payload}).encode("utf-8"))
    body_raw = resp["Payload"].read().decode("utf-8")
    try:
        data = json.loads(body_raw)
        inner = data.get("body")
        if isinstance(inner, str):
            inner = json.loads(inner)
        return inner.get("message") or _draft_direct(student, flags)
    except Exception:
        return _draft_direct(student, flags)


def _draft_direct(student: Dict[str, Any], flags: List[str]) -> str:
    try:
        import boto3  # type: ignore
    except Exception:
        return _template(student, flags)
    model_id = os.getenv("BEDROCK_MODEL_ID", "us.anthropic.claude-sonnet-4-20250514-v1:0")
    region = os.getenv("AWS_REGION", "us-east-1")
    br = boto3.client("bedrock-runtime", region_name=region)
    prompt = (
        "You are an empathetic instructor. Draft a concise, supportive Slack DM (<= 3 sentences) "
        "to a learner based on the following context. Avoid shaming; be specific, offer help, and keep a warm tone.\n\n"
        f"Learner: {student['studentName']}\n"
        f"Signals: {', '.join(flags)}\n"
        f"Metrics: {student.get('metrics', {})}\n"
    )
    resp = br.converse(
        modelId=model_id,
        messages=[{"role": "user", "content": [{"type": "text", "text": prompt}]}],
        inferenceConfig={"maxTokens": 300, "temperature": 0.3},
    )
    try:
        return resp["output"]["message"]["content"][0]["text"]
    except Exception:
        return _template(student, flags)


def _template(student: Dict[str, Any], flags: List[str]) -> str:
    reasons = ", ".join(flags)
    return (
        f"Hi {student['studentName']}, I noticed a few signals this week ({reasons}). "
        f"How are you feeling about the material? Anything I can clarify or help with?"
    )


def _queue_notification(api_base: str, student: Dict[str, Any], message: str) -> None:
    import os as _os
    import json as _json
    import urllib.request as _url

    endpoint = api_base.rstrip("/") + "/queue"
    data = _json.dumps(
        {
            "studentId": student["studentId"],
            "studentName": student["studentName"],
            "slackUserId": student["slackUserId"],
            "message": message,
        }
    ).encode("utf-8")
    req = _url.Request(endpoint, data=data, headers={"Content-Type": "application/json"}, method="POST")
    with _url.urlopen(req, timeout=10) as resp:  # noqa: S310
        _ = resp.read()


def _queue_with_flags(api_base: str, current_by_student: Dict[str, Dict[str, Any]], flags_by_student: Dict[str, List[str]]) -> int:
    queued = 0
    for sid, flags in flags_by_student.items():
        student = current_by_student.get(sid)
        if not student:
            continue
        msg = _invoke_draft_lambda(student, flags)
        try:
            _queue_notification(api_base, student, msg)
            queued += 1
        except Exception:
            pass
    return queued


def _llm_decide_and_queue(api_base: str, students: List[Dict[str, Any]]) -> set[str]:
    try:
        import boto3  # type: ignore
    except Exception:
        return set()
    model_id = os.getenv("BEDROCK_MODEL_ID", "us.anthropic.claude-sonnet-4-20250514-v1:0")
    region = os.getenv("AWS_REGION", "us-east-1")
    br = boto3.client("bedrock-runtime", region_name=region)
    # Keep prompt compact; pass rows as JSON to the model
    import json as _json
    prompt = (
        "You are an instructional coach. Given a JSON array of student rows with metrics, "
        "choose students who should receive a supportive DM this week and write a short message for each. "
        "Be selective, avoid duplicates. Return strictly JSON: {\"decisions\":[{\"studentId\":...,\"message\":...}]}\n\n"
        f"rows: {_json.dumps(students)[:12000]}\n"
    )
    resp = br.converse(
        modelId=model_id,
        messages=[{"role": "user", "content": [{"text": prompt}]}],
        inferenceConfig={"maxTokens": 500, "temperature": 0.2},
    )
    try:
        text = resp["output"]["message"]["content"][0]["text"]
        out = _json.loads(text)
        processed: set[str] = set()
        for d in out.get("decisions", []):
            sid = str(d.get("studentId"))
            msg = str(d.get("message", "")).strip()
            if not sid or not msg:
                continue
            st = next((s for s in students if str(s.get("studentId")) == sid), None)
            if st:
                try:
                    _queue_notification(api_base, st, msg)
                    processed.add(sid)
                except Exception:
                    pass
        return processed
    except Exception:
        return set()

def handler(event, context):
    bucket = os.environ.get("BUCKET_NAME")
    table = os.environ.get("TABLE_NAME")  # reserved for future use
    api_base = os.environ.get("QUEUE_API_BASE", "https://REPLACE_WITH_API")
    use_llm_decide = os.environ.get("USE_LLM_DECIDE", "false").lower() in {"1", "true", "yes"}

    # S3 Put event
    records = event.get("Records", [])
    if not records:
        return {"statusCode": 200, "body": json.dumps({"ok": True, "note": "no records"})}

    # Collect latest and previous files if available by key convention YYYY-MM-DD.xlsx
    s3_keys = [r["s3"]["object"]["key"] for r in records]
    key = s3_keys[0]

    # Load current upload
    current_records = _read_excel_from_s3(bucket, key)
    current_by_student = {r["studentId"]: r for r in current_records}

    # Optionally find previous file in same prefix
    prev_by_student: Dict[str, Dict[str, Any]] = {}
    try:
        import re
        import boto3

        s3 = boto3.client("s3")
        prefix = key.rsplit("/", 1)[0] + "/" if "/" in key else ""
        listed = s3.list_objects_v2(Bucket=bucket, Prefix=prefix)
        candidates = [it["Key"] for it in listed.get("Contents", []) if it["Key"].endswith(".xlsx")]
        candidates = sorted(candidates)
        if key in candidates:
            idx = candidates.index(key)
            if idx > 0:
                prev_key = candidates[idx - 1]
                prev_records = _read_excel_from_s3(bucket, prev_key)
                prev_by_student = {r["studentId"]: r for r in prev_records}
    except Exception:
        prev_by_student = {}

    if use_llm_decide:
        try:
            processed_sids = _llm_decide_and_queue(api_base, list(current_by_student.values()))
            queued = len(processed_sids)
            # Also queue any remaining flagged students using rule-based logic
            deltas_by_student = _compute_deltas(prev_by_student, current_by_student)
            flags_by_student = _decide_flags(current_by_student, deltas_by_student)
            if processed_sids:
                flags_by_student = {sid: f for sid, f in flags_by_student.items() if sid not in processed_sids}
            if flags_by_student:
                queued += _queue_with_flags(api_base, current_by_student, flags_by_student)
        except Exception:
            # Fallback to rule-based if LLM decision fails
            deltas_by_student = _compute_deltas(prev_by_student, current_by_student)
            flags_by_student = _decide_flags(current_by_student, deltas_by_student)
            queued = _queue_with_flags(api_base, current_by_student, flags_by_student)
    else:
        deltas_by_student = _compute_deltas(prev_by_student, current_by_student)
        flags_by_student = _decide_flags(current_by_student, deltas_by_student)
        queued = _queue_with_flags(api_base, current_by_student, flags_by_student)

    return {"statusCode": 200, "body": json.dumps({"ok": True, "queued": int(queued)})}


