import json
import io
from typing import Any, Dict, List


def _read_excel_bytes(content: bytes) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook  # type: ignore

    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") if h is not None else "" for h in rows[0]]
    required = {"student_id", "student_name", "slack_user_id"}
    if not required.issubset(set(headers)):
        missing = sorted(required - set(headers))
        raise ValueError(f"Missing column(s): {missing}")

    id_cols = {"student_id", "student_name", "slack_user_id", "last_active"}
    metric_idxs = [i for i, h in enumerate(headers) if h and h not in id_cols]

    records: List[Dict[str, Any]] = []
    for r in rows[1:]:
        row = list(r) if r is not None else []
        vals = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        metrics: Dict[str, float] = {}
        for i in metric_idxs:
            if i < len(row) and row[i] is not None and str(row[i]).strip() != "":
                try:
                    metrics[headers[i]] = float(row[i])
                except Exception:
                    continue
        last_active_val = vals.get("last_active")
        last_iso = str(last_active_val) if last_active_val not in (None, "") else None
        records.append(
            {
                "studentId": str(vals.get("student_id", "")),
                "studentName": str(vals.get("student_name", "")),
                "slackUserId": str(vals.get("slack_user_id", "")),
                "metrics": metrics,
                "lastActiveIso": last_iso,
            }
        )
    return records


def handler(event, context):
    body = event.get("body")
    if isinstance(body, str):
        body = json.loads(body)

    # Expect { bucket, key } to read from S3, or { contentBase64 }
    bucket = None
    key = None
    content_b64 = None
    if isinstance(body, dict):
        bucket = body.get("bucket")
        key = body.get("key")
        content_b64 = body.get("contentBase64")

    try:
        if content_b64:
            import base64

            content = base64.b64decode(content_b64)
            records = _read_excel_bytes(content)
        elif bucket and key:
            import boto3  # type: ignore

            s3 = boto3.client("s3")
            obj = s3.get_object(Bucket=bucket, Key=key)
            content = obj["Body"].read()
            records = _read_excel_bytes(content)
        else:
            raise ValueError("Provide contentBase64 or bucket+key")
    except Exception as e:
        return {
            "statusCode": 400,
            "headers": {"content-type": "application/json"},
            "body": json.dumps({"error": str(e)}),
        }

    return {
        "statusCode": 200,
        "headers": {"content-type": "application/json"},
        "body": json.dumps({"records": records}),
    }


